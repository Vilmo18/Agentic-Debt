'''
Excel storage for Budget Tracker using openpyxl.
Creates/loads/saves a workbook with sheets: Transactions, Categories, Settings.
'''
import os
from typing import List, Iterable
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from models import Transaction, BudgetModel


class ExcelStorage:
    def __init__(self, filepath: str):
        self.filepath = filepath
        # Ensure directory exists
        directory = os.path.dirname(os.path.abspath(self.filepath))
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        # Initialize file if missing
        if not os.path.exists(self.filepath):
            self._initialize_empty_file()

    def _initialize_empty_file(self) -> None:
        wb = Workbook()
        ws_trans = wb.active
        ws_trans.title = "Transactions"
        self._write_transactions_header(ws_trans)

        ws_cat = wb.create_sheet("Categories")
        ws_cat.append(["Name"])
        for name in ["General", "Food", "Transport", "Housing", "Utilities", "Entertainment", "Healthcare", "Income:Salary", "Income:Other", "Savings"]:
            ws_cat.append([name])

        ws_settings = wb.create_sheet("Settings")
        ws_settings.append(["Key", "Value"])
        ws_settings.append(["savings_goal", 0.0])

        self._autosize_columns(ws_trans)
        self._autosize_columns(ws_cat)
        self._autosize_columns(ws_settings)

        wb.save(self.filepath)

    @staticmethod
    def _write_transactions_header(ws: Worksheet) -> None:
        ws.append(["ID", "Date", "Type", "Category", "Description", "Amount"])

    @staticmethod
    def _autosize_columns(ws: Worksheet) -> None:
        # Compute max length in each column
        widths = {}
        for row in ws.iter_rows(values_only=True):
            for idx, value in enumerate(row, start=1):
                txt = "" if value is None else str(value)
                widths[idx] = max(widths.get(idx, 0), len(txt))
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 40)

    def load_model(self) -> BudgetModel:
        if not os.path.exists(self.filepath):
            self._initialize_empty_file()

        wb = load_workbook(self.filepath, data_only=True)
        # Transactions
        trans_sheet = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active
        transactions: List[Transaction] = []
        first = True
        for row in trans_sheet.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or row[0] is None:
                continue
            try:
                tx = Transaction.from_row(row)
                transactions.append(tx)
            except Exception:
                # Skip malformed rows
                continue

        # Categories
        categories: List[str] = []
        if "Categories" in wb.sheetnames:
            cat_sheet = wb["Categories"]
            first = True
            for row in cat_sheet.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if row and row[0]:
                    categories.append(str(row[0]).strip())

        # Settings
        savings_goal = 0.0
        if "Settings" in wb.sheetnames:
            set_sheet = wb["Settings"]
            first = True
            for row in set_sheet.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if row and row[0] == "savings_goal":
                    try:
                        savings_goal = float(row[1])
                    except Exception:
                        savings_goal = 0.0
                    break

        model = BudgetModel()
        model.load(transactions=transactions, categories=categories, savings_goal=savings_goal)
        return model

    def save_model(self, model: BudgetModel) -> None:
        # Create a fresh workbook and write all content
        wb = Workbook()
        ws_trans = wb.active
        ws_trans.title = "Transactions"
        self._write_transactions_header(ws_trans)
        for t in model.transactions:
            row = list(t.to_row())
            # Ensure Amount stored as positive number; sign inferred by Type
            row[5] = float(abs(row[5]))
            ws_trans.append(row)

        ws_cat = wb.create_sheet("Categories")
        ws_cat.append(["Name"])
        for name in model.categories:
            ws_cat.append([name])

        ws_settings = wb.create_sheet("Settings")
        ws_settings.append(["Key", "Value"])
        ws_settings.append(["savings_goal", float(model.savings_goal)])

        # Formatting and sizes
        self._autosize_columns(ws_trans)
        self._autosize_columns(ws_cat)
        self._autosize_columns(ws_settings)

        wb.save(self.filepath)